from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import workbook,load_workbook
wb = load_workbook("googlesearchresults.xlsx")
inp=input()
sh=wb.create_sheet(inp)
sh.append(["NAme","moreinfo"])
inp=inp.replace(" ",'+')
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)

driver = webdriver.Chrome(options=chrome_options)
driver.implicitly_wait(5)
driver.get('https://www.google.com/search?q='+inp+"&start")
search=driver.find_elements(By.XPATH,"//div[@class='MjjYud']")
for ele in search:
    try:
        sh.append([ele.find_element(By.XPATH,".//h3").text.replace("-" ,""), ele.find_element(By.XPATH,".//a").get_attribute("href")])
    except:
        pass
wb.save("googlesearchresults.xlsx")
